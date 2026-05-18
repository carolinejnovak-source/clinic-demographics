import time
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
import openpyxl
import io
import flexpolyline
import requests
import math
import os
import re
import json
import threading
from auth import login_required, check_credentials
from ring_ads import compute_ring_ads
from patient_heatmap import get_heatmap_points
from clinic_performance import get_clinic_performance, get_campaigns_list

# ── CTR% data from test campaigns (35 clinics) ──────────────────────────────
CTR_DATA = {
    'Midtown':10.8,'FiDi':11.0,'Brooklyn':11.0,'Upper East Side':11.0,
    'Harrison':11.0,'Clifton':10.8,'Edgewater':10.9,'Paramus':11.0,
    'Woodland':10.8,'Hoboken':10.7,'West Orange':10.9,'Astoria':10.8,
    'Forest Hills':10.7,'Brighton Beach':13.2,'Bronx':11.1,
    'Morris County':11.2,'Morristown':11.2,'Yonkers':11.8,
    'Woodbridge':10.3,'West Islip':13.1,'Hartsdale':12.9,
    'Staten Island':14.1,'Port Jefferson':11.4,'San Diego':8.5,
    'National City':9.1,'Cedar Park':14.5,'Fort Worth':11.2,
    'Huntington Beach':9.7,'Newport Beach':9.5,'Arlington':10.8,
    'Princeton':8.3,'Kyle':10.9,'Hamden':8.3,'Temecula':16.7,'Addison':4.0,
}

# ── Scoring helpers ──────────────────────────────────────────────────────────
import math as _math

# Reference distribution built from feature_matrix for normalisation
_SCORE_REFS = {
    'hhi_mean': 11.75, 'hhi_std': 0.38,   # log(median HHI)
    'dpop_mean': 11.60, 'dpop_std': 0.80,  # log(daytime pop)
    'ctr_mean': 2.40,  'ctr_std': 0.35,    # log(CTR%)
}

def _zscore(val, mean, std):
    if val is None or std == 0: return None
    return (val - mean) / std

def compute_base_score(median_hhi, daytime_pop):
    """HHI 70% + Daytime Pop 30% — greenfield-safe"""
    z_hhi  = _zscore(_math.log(median_hhi)  if median_hhi  and median_hhi>0  else None, _SCORE_REFS['hhi_mean'],  _SCORE_REFS['hhi_std'])
    z_dpop = _zscore(_math.log(daytime_pop) if daytime_pop and daytime_pop>0 else None, _SCORE_REFS['dpop_mean'], _SCORE_REFS['dpop_std'])
    if z_hhi is None and z_dpop is None: return None
    z = (0.70*(z_hhi or 0) + 0.30*(z_dpop or 0))
    return max(0, min(100, round(50 + z*15)))

def compute_enhanced_score(median_hhi, daytime_pop, ctr_pct):
    """HHI 50% + CTR 30% + Daytime Pop 20% — requires test campaign data"""
    z_hhi  = _zscore(_math.log(median_hhi)  if median_hhi  and median_hhi>0  else None, _SCORE_REFS['hhi_mean'],  _SCORE_REFS['hhi_std'])
    z_dpop = _zscore(_math.log(daytime_pop) if daytime_pop and daytime_pop>0 else None, _SCORE_REFS['dpop_mean'], _SCORE_REFS['dpop_std'])
    z_ctr  = _zscore(_math.log(ctr_pct)     if ctr_pct     and ctr_pct>0     else None, _SCORE_REFS['ctr_mean'],  _SCORE_REFS['ctr_std'])
    if z_hhi is None or z_ctr is None: return None
    z = (0.50*(z_hhi or 0) + 0.30*(z_ctr or 0) + 0.20*(z_dpop or 0))
    return max(0, min(100, round(50 + z*15)))



app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "vtc-clinic-demo-secret-2024")
app.config["SESSION_COOKIE_PATH"] = "/"
app.config["SESSION_COOKIE_NAME"] = "cd_session"
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if check_credentials(username, password):
            session["logged_in"] = True
            next_url = request.args.get("next", "/clinic-demographics/")
            if not next_url.startswith("/clinic-demographics"):
                next_url = "/clinic-demographics" + next_url
            return redirect(next_url)
        flash("Invalid username or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

SITES_FILE = '/opt/mikala-apps/clinic-demographics/vip_sites.json'
STATUS_COLORS = {
    'open': 'blue',
    'pending opening day': 'purple',
    'possible new sites': 'orange',
}

def load_sites():
    try:
        with open(SITES_FILE) as f:
            return json.load(f)
    except:
        return {"territories": {}, "clinics": []}

def get_clinics():
    return load_sites().get("clinics", [])

# Keep backward compat
CLINICS = get_clinics()

STATE_FIPS = {
    "AL":"01","AK":"02","AZ":"04","AR":"05","CA":"06","CO":"08","CT":"09",
    "DE":"10","DC":"11","FL":"12","GA":"13","HI":"15","ID":"16","IL":"17",
    "IN":"18","IA":"19","KS":"20","KY":"21","LA":"22","ME":"23","MD":"24",
    "MA":"25","MI":"26","MN":"27","MS":"28","MO":"29","MT":"30","NE":"31",
    "NV":"32","NH":"33","NJ":"34","NM":"35","NY":"36","NC":"37","ND":"38",
    "OH":"39","OK":"40","OR":"41","PA":"42","RI":"44","SC":"45","SD":"46",
    "TN":"47","TX":"48","UT":"49","VT":"50","VA":"51","WA":"53","WV":"54",
    "WI":"55","WY":"56",
}

_cache = {}
_DISK_CACHE_FILE = os.path.join(os.path.dirname(__file__), 'geo_iso_cache.json')

def _load_disk_cache():
    global _cache
    try:
        with open(_DISK_CACHE_FILE) as f:
            _cache = json.load(f)
        print(f"Loaded {len(_cache)} cached entries from disk")
    except Exception:
        _cache = {}

def _save_disk_cache():
    try:
        # Only save geo + iso keys (not raw population data - too large)
        to_save = {k: v for k, v in _cache.items() if k.startswith('geo:') or k.startswith('iso:')}
        with open(_DISK_CACHE_FILE, 'w') as f:
            json.dump(to_save, f)
    except Exception as e:
        print(f"Cache save error: {e}")

_load_disk_cache()

def _evict_clinic_cache(address):
    """Remove all cache entries related to a specific clinic address."""
    # Geocode key
    geo_key = f"geo:{address}"
    coords = _cache.pop(geo_key, None)
    if coords:
        lat, lon = coords
        iso_key = f"iso:{lat:.4f},{lon:.4f}"
        iso = _cache.pop(iso_key, None)
        # Pop ring population keys
        for ring_key in list(_cache.keys()):
            if ring_key.startswith("pop_ring:"):
                _cache.pop(ring_key, None)
    # County/zip based keys - harder to target, leave them (they're shared across clinics)

def _smart_cache_refresh(old_clinics, new_clinics):
    """Only evict cache for clinics that are new or have changed addresses/status."""
    old_addresses = {c["address"]: c for c in old_clinics}
    new_addresses = {c["address"]: c for c in new_clinics}
    # Evict removed or changed clinics
    for addr, clinic in old_addresses.items():
        if addr not in new_addresses:
            _evict_clinic_cache(addr)
    # Evict new clinics (they have no cache yet, but preloader needs to know to load them)
    new_only = [c for c in new_clinics if c["address"] not in old_addresses]
    return new_only  # return list of new clinics to preload
_preload_status = {"done": 0, "total": len(CLINICS), "complete": False}


# ── Helpers ──────────────────────────────────────────────────────

def extract_zip(address):
    m = re.search(r'\b(\d{5})\b', address)
    return m.group(1) if m else None

def extract_state(address):
    m = re.search(r',\s*([A-Z]{2})\s+\d{5}', address)
    return m.group(1) if m else None

def geocode_address(address):
    key = f"geo:{address}"
    if key in _cache:
        return _cache[key]
    try:
        r = requests.get(
            "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress",
            params={"address": address, "benchmark": "2020", "format": "json"},
            timeout=10
        )
        matches = r.json().get("result", {}).get("addressMatches", [])
        if matches:
            c = matches[0]["coordinates"]
            result = (float(c["y"]), float(c["x"]))
            _cache[key] = result
            _save_disk_cache()
            return result
    except Exception as e:
        print(f"Geocode error for {address}: {e}")
    return None

HERE_API_KEY = "WYqJdDCGYmTRFQT5zxM_aUNTb8XH0_MwezLgqvRkCbE"
CENSUS_API_KEY = "39d91e3ea57b794ee42f0e60b4548835eb21293b"

def _census_get(url, params, timeout=15, retries=3):
    """Census API call with key injection, retry on 429."""
    import time
    if CENSUS_API_KEY:
        params = dict(params, key=CENSUS_API_KEY)
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, timeout=timeout)
            if r.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            return r
        except Exception as e:
            if attempt == retries - 1:
                raise
            time.sleep(1)
    return None
HERE_DEPARTURE = "2026-03-10T10:00:00"  # Tuesday 10am — matches Buxton methodology

def _here_polygon_to_geojson(encoded, minutes):
    """Decode HERE flexible polyline and return a GeoJSON feature."""
    coords = flexpolyline.decode(encoded)
    # coords are (lat, lon) tuples — GeoJSON needs [lon, lat]
    ring = [[lon, lat] for lat, lon in coords]
    if ring[0] != ring[-1]:
        ring.append(ring[0])
    return {
        "type": "Feature",
        "geometry": {"type": "Polygon", "coordinates": [ring]},
        "properties": {"contour": minutes, "metric": "time"}
    }

def get_isochrone(lat, lon):
    key = f"iso:{lat:.4f},{lon:.4f}"
    if key in _cache:
        return _cache[key]
    try:
        url = (
            f"https://isoline.router.hereapi.com/v8/isolines"
            f"?transportMode=car"
            f"&origin={lat},{lon}"
            f"&range%5Btype%5D=time"
            f"&range%5Bvalues%5D=600,1200"
            f"&departureTime={HERE_DEPARTURE}"
            f"&apikey={HERE_API_KEY}"
        )
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            data = r.json()
            features = []
            for isoline in data.get("isolines", []):
                minutes = isoline["range"]["value"] // 60
                for poly in isoline.get("polygons", []):
                    features.append(_here_polygon_to_geojson(poly["outer"], minutes))
            geojson = {"type": "FeatureCollection", "features": features}
            _cache[key] = geojson
            return geojson
    except Exception as e:
        print(f"HERE isochrone error for ({lat},{lon}): {e}")
    return None

def geocode_state(lat, lon):
    """Return 2-letter state abbreviation for a lat/lon using Census geocoder."""
    key = f"state:{lat:.3f},{lon:.3f}"
    if key in _cache:
        return _cache[key]
    try:
        r = requests.get(
            "https://geocoding.geo.census.gov/geocoder/geographies/coordinates",
            params={"x": lon, "y": lat, "benchmark": "Public_AR_Current",
                    "vintage": "Current_Current", "layers": "States", "format": "json"},
            timeout=10)
        result = r.json()
        states = result["result"]["geographies"].get("States", [])
        abbr = states[0].get("STUSAB", "") if states else ""
        _cache[key] = abbr
        return abbr
    except:
        return ""

def get_county_fips(lat, lon):
    key = f"county:{lat:.4f},{lon:.4f}"
    if key in _cache:
        return _cache[key]
    try:
        r = requests.get(
            "https://geocoding.geo.census.gov/geocoder/geographies/coordinates",
            params={"x": lon, "y": lat,
                    "benchmark": "Public_AR_Current",
                    "vintage": "Current_Current",
                    "layers": "Counties", "format": "json"},
            timeout=10
        )
        counties = r.json().get("result", {}).get("geographies", {}).get("Counties", [])
        if counties:
            c = counties[0]
            result = (c.get("STATE", ""), c.get("COUNTY", ""))
            _cache[key] = result
            return result
    except:
        pass
    return (None, None)

def get_census_acs(state_fips, county_fips):
    key = f"acs:{state_fips},{county_fips}"
    if key in _cache:
        return _cache[key]
    result = {"population": None, "median_income": None, "error": None}
    try:
        female_vars = ",".join(f"B01001_{n:03d}E" for n in range(37, 50))
        all_vars = f"B01003_001E,B19013_001E,{female_vars}"
        r = _census_get(
            "https://api.census.gov/data/2022/acs/acs5",
            params={"get": all_vars,
                    "for": f"county:{county_fips}",
                    "in": f"state:{state_fips}"}
        )
        data = r.json()
        if len(data) >= 2:
            d = dict(zip(data[0], data[1]))
            pop = int(d.get("B01003_001E", 0) or 0)
            inc = int(d.get("B19013_001E", 0) or 0)
            f35 = sum(int(d.get(f"B01001_{n:03d}E", 0) or 0) for n in range(37, 50))
            result["population"] = pop if pop > 0 else None
            result["median_income"] = inc if inc > 0 else None
            result["female_35plus"] = f35 if f35 > 0 else None
    except Exception as e:
        result["error"] = str(e)
    _cache[key] = result
    return result


def get_zip_income(zipcode):
    """Median household income for a ZIP code (Census ACS 5-year, ZCTA level)."""
    key = f"zip_income:{zipcode}"
    if key in _cache:
        return _cache[key]
    result = None
    try:
        r = _census_get(
            "https://api.census.gov/data/2022/acs/acs5",
            params={"get": "B19013_001E",
                    "for": f"zip code tabulation area:{zipcode}"}
        )
        if r and r.ok:
            data = r.json()
            if len(data) >= 2:
                val = int(data[1][0] or 0)
                result = val if val > 0 else None
    except Exception as e:
        print(f"ZIP income error {zipcode}: {e}")
    _cache[key] = result
    return result

def get_sahie(state_fips, county_fips):
    key = f"sahie:{state_fips},{county_fips}"
    if key in _cache:
        return _cache[key]
    result = {"insured_pct": None, "error": None}
    try:
        r = requests.get(
            "https://api.census.gov/data/timeseries/healthins/sahie",
            params={"get": "PCTIC_PT",
                    "for": f"county:{county_fips}",
                    "in": f"state:{state_fips}",
                    "time": "2022",
                    "key": CENSUS_API_KEY},
            timeout=8
        )
        data = r.json()
        if len(data) >= 2:
            d = dict(zip(data[0], data[1]))
            pct = float(d.get("PCTIC_PT", 0) or 0)
            result["insured_pct"] = round(pct, 1) if pct > 0 else None
    except Exception as e:
        result["error"] = str(e)
    _cache[key] = result
    return result

# CMS Medicare Physician & Other Practitioners - by Provider and Service
CMS_DATASET = "https://data.cms.gov/data-api/v1/dataset/92396110-2aed-4d63-a6a2-5d6207d46a29/data"
CMS_GEO_DATASET = "https://data.cms.gov/data-api/v1/dataset/6fea9d79-0129-4e4c-b1b8-23cd86a4f435/data"
VTC_CPT_CODES = ["36475", "36465", "36466"]

def get_cms(zipcode, state_fips=None):
    """Fetch CPT volume + reimbursement rates. Uses state-level for volume/rates, ZIP for competitors."""
    key = f"cms2:{state_fips or zipcode}"
    if key in _cache:
        return _cache[key]
    result = {
        "cpt36475_volume": 0, "cpt36465_volume": 0, "cpt36466_volume": 0,
        "cpt_total_volume": 0,
        "medicare_rate_36475": None, "medicare_rate_36465": None, "medicare_rate_36466": None,
        "competitors": [],
        "error": None
    }
    def _fetch_cpt_state(cpt):
        """State-level volume + rates."""
        try:
            r = requests.get(CMS_GEO_DATASET,
                params={"filter[HCPCS_Cd]": cpt,
                        "filter[Rndrng_Prvdr_Geo_Cd]": state_fips or "",
                        "filter[Rndrng_Prvdr_Geo_Lvl]": "State", "size": 10},
                timeout=15)
            if r.ok and isinstance(r.json(), list):
                return cpt, r.json()
        except:
            pass
        return cpt, []

    def _fetch_cpt_zip(cpt):
        """ZIP-level for competitor detection (36475 only)."""
        try:
            r = requests.get(CMS_DATASET,
                params={"filter[Rndrng_Prvdr_Zip5]": zipcode,
                        "filter[HCPCS_Cd]": cpt, "size": 500},
                timeout=15)
            if r.ok and isinstance(r.json(), list):
                return cpt, r.json()
        except:
            pass
        return cpt, []

    try:
        from concurrent.futures import ThreadPoolExecutor as _TPE
        with _TPE(max_workers=4) as ex:
            state_results = dict(ex.map(lambda c: _fetch_cpt_state(c), VTC_CPT_CODES))
            f_zip = ex.submit(lambda: _fetch_cpt_zip("36475"))
        _, zip_rows = f_zip.result()

        for cpt, rows in state_results.items():
            # Sum across office + facility place of service
            vol = sum(int(row.get("Tot_Srvcs", 0) or 0) for row in rows)
            result[f"cpt{cpt}_volume"] = vol
            result["cpt_total_volume"] += vol
            total_pay = sum(
                float(row.get("Avg_Mdcr_Pymt_Amt", 0) or 0) * int(row.get("Tot_Srvcs", 0) or 0)
                for row in rows)
            if vol > 0:
                result[f"medicare_rate_{cpt}"] = round(total_pay / vol, 2)

        # Competitors from ZIP-level data
        for row in zip_rows:
            name = f"{row.get('Rndrng_Prvdr_First_Name','')} {row.get('Rndrng_Prvdr_Last_Org_Name','')}".strip()
            result["competitors"].append({
                "name": name,
                "type": row.get("Rndrng_Prvdr_Type",""),
                "city": row.get("Rndrng_Prvdr_City",""),
                "state": row.get("Rndrng_Prvdr_State_Abrvtn",""),
                "zip": row.get("Rndrng_Prvdr_Zip5",""),
                "npi": row.get("Rndrng_NPI",""),
                "volume": int(row.get("Tot_Srvcs", 0) or 0),
                "medicare_participating": row.get("Rndrng_Prvdr_Mdcr_Prtcptg_Ind","") == "Y",
            })
    except Exception as e:
        result["error"] = str(e)
    _cache[key] = result
    return result


def get_cms_state(state_abbr):
    """Fetch all CPT 36475 providers in a state for competitor radius search."""
    key = f"cms_state:{state_abbr}"
    if key in _cache:
        return _cache[key]
    try:
        r = requests.get(CMS_DATASET,
            params={"filter[Rndrng_Prvdr_State_Abrvtn]": state_abbr,
                    "filter[HCPCS_Cd]": "36475", "size": 2000},
            timeout=30)
        if r.ok and isinstance(r.json(), list):
            result = r.json()
            _cache[key] = result
            return result
    except Exception as e:
        print(f"CMS state fetch error: {e}")
    return []


def get_competitors_near(lat, lon, state_abbr, radius_km=40):
    """Find providers billing CPT 36475 within radius_km of lat/lon."""
    import math
    providers = get_cms_state(state_abbr)
    nearby = []
    for p in providers:
        pzip = p.get("Rndrng_Prvdr_Zip5","")
        # Quick filter: use ZIP centroid from geocoder cache if available
        pkey = f"zip:{pzip}"
        if pkey not in _cache:
            continue
        plat, plon = _cache[pkey]
        # Haversine distance
        R = 6371
        dlat = math.radians(plat - lat)
        dlon = math.radians(plon - lon)
        a = math.sin(dlat/2)**2 + math.cos(math.radians(lat)) * math.cos(math.radians(plat)) * math.sin(dlon/2)**2
        dist = R * 2 * math.asin(math.sqrt(a))
        if dist <= radius_km:
            name = f"{p.get('Rndrng_Prvdr_First_Name','')} {p.get('Rndrng_Prvdr_Last_Org_Name','')}".strip()
            nearby.append({
                "name": name,
                "type": p.get("Rndrng_Prvdr_Type",""),
                "city": p.get("Rndrng_Prvdr_City",""),
                "state": p.get("Rndrng_Prvdr_State_Abrvtn",""),
                "zip": pzip,
                "npi": p.get("Rndrng_NPI",""),
                "volume": int(p.get("Tot_Srvcs", 0) or 0),
                "medicare_participating": p.get("Rndrng_Prvdr_Mdcr_Prtcptg_Ind","") == "Y",
                "distance_km": round(dist, 1),
            })
    nearby.sort(key=lambda x: x["distance_km"])
    return nearby


def get_population_in_ring(isochrone_geojson, ring_minutes):
    """
    Returns population statistics for the area within an isochrone ring.
    Uses Census block-group data with Shapely polygon intersection.
    """
    from collections import defaultdict
    try:
        from shapely.geometry import shape, Point
    except ImportError:
        return {"total_pop": None, "female_35plus": None, "median_income": None}

    if not isochrone_geojson:
        return {"total_pop": None, "female_35plus": None, "median_income": None}

    cache_key = f"pop_ring:{ring_minutes}:{hash(str(isochrone_geojson))}"
    if cache_key in _cache:
        return _cache[cache_key]

    result = {"total_pop": None, "female_35plus": None, "median_income": None}

    try:
        features = isochrone_geojson.get("features", [])
        # Find the polygon for this ring (contour == ring_minutes)
        poly_geom = None
        for f in features:
            if f.get("properties", {}).get("contour") == ring_minutes:
                poly_geom = f["geometry"]
                break
        if not poly_geom:
            _cache[cache_key] = result
            return result

        ring_shape = shape(poly_geom)
        coords_list = poly_geom["coordinates"][0]
        lons = [c[0] for c in coords_list]
        lats = [c[1] for c in coords_list]
        bbox = f"{min(lons)},{min(lats)},{max(lons)},{max(lats)}"

        # Get block groups from TIGERweb that intersect the bounding box
        r = requests.get(
            "https://tigerweb.geo.census.gov/arcgis/rest/services/TIGERweb/tigerWMS_ACS2022/MapServer/8/query",
            params={"geometry": bbox, "geometryType": "esriGeometryEnvelope",
                    "inSR": "4326", "spatialRel": "esriSpatialRelIntersects",
                    "outFields": "GEOID,STATE,COUNTY,TRACT,BLKGRP,CENTLAT,CENTLON",
                    "returnGeometry": "false", "f": "json"},
            timeout=20)
        if not r.ok:
            _cache[cache_key] = result
            return result

        features_bg = r.json().get("features", [])
        if not features_bg:
            _cache[cache_key] = result
            return result

        # Filter block groups whose centroid is within the ring polygon
        by_state_county = defaultdict(list)
        for feat in features_bg:
            attr = feat.get("attributes", {})
            clat = attr.get("CENTLAT")
            clon = attr.get("CENTLON")
            if clat is None or clon is None:
                # No centroid — include it anyway (conservative)
                pass
            else:
                try:
                    pt = Point(float(clon), float(clat))
                    if not ring_shape.contains(pt):
                        continue
                except:
                    pass
            sc = (str(attr.get("STATE", "")).zfill(2), str(attr.get("COUNTY", "")).zfill(3))
            by_state_county[sc].append((str(attr.get("TRACT", "")).zfill(6), str(attr.get("BLKGRP", ""))))

        if not by_state_county:
            _cache[cache_key] = result
            return result

        total_pop = 0
        female_35plus = 0
        income_vals = []
        female_vars = ",".join(f"B01001_{n:03d}E" for n in range(37, 50))
        vars_needed = f"B01003_001E,{female_vars},B19013_001E"

        for (state, county), tracts in by_state_county.items():
            params = {"get": vars_needed, "for": "block group:*",
                      "in": f"state:{state} county:{county}"}
            rc = _census_get("https://api.census.gov/data/2022/acs/acs5", params)
            if not rc or not rc.ok:
                continue
            try:
                rows = rc.json()
                if not isinstance(rows, list) or len(rows) < 2:
                    continue
                headers = rows[0]
                idx = {h: i for i, h in enumerate(headers)}
                tract_idx = idx.get("tract")
                bg_idx = idx.get("block group")
                wanted = {(t, b) for t, b in tracts}
                for row in rows[1:]:
                    if tract_idx is None or bg_idx is None:
                        continue
                    tract = str(row[tract_idx]).zfill(6)
                    bg = str(row[bg_idx])
                    if (tract, bg) not in wanted:
                        continue
                    def si(v):
                        try: return max(0, int(v or 0))
                        except: return 0
                    pop = si(row[idx.get("B01003_001E", -1)])
                    f35 = sum(si(row[idx.get(f"B01001_{n:03d}E", -1)]) for n in range(37, 50))
                    inc_raw = row[idx.get("B19013_001E", -1)]
                    total_pop += pop
                    female_35plus += f35
                    try:
                        iv = int(inc_raw)
                        if iv > 0 and pop > 0:
                            income_vals.append((iv, pop))
                    except:
                        pass
            except Exception as e:
                print(f"ACS BG parse error {state}/{county}: {e}")
                continue

        result["total_pop"] = total_pop if total_pop > 0 else None
        result["female_35plus"] = female_35plus if female_35plus > 0 else None
        if income_vals:
            tw = sum(w for _, w in income_vals)
            result["median_income"] = round(sum(v * w for v, w in income_vals) / tw) if tw > 0 else None

    except Exception as e:
        print(f"Population ring error: {e}")

    _cache[cache_key] = result
    return result



# ── Background preloader ─────────────────────────────────────────

def _preload_worker():
    """Geocode all clinics, fetch isochrones, and preload all demographics so clicks are instant."""
    from concurrent.futures import ThreadPoolExecutor
    def load_one(clinic):
        try:
            address = clinic["address"]
            coords = geocode_address(address)
            if not coords:
                return
            lat, lon = coords
            # Isochrone
            iso = get_isochrone(lat, lon)
            # Demographics
            zipcode = extract_zip(address)
            state_fips, county_fips = get_county_fips(lat, lon)
            if state_fips:
                get_census_acs(state_fips, county_fips)
                get_sahie(state_fips, county_fips)
            if zipcode:
                get_cms(zipcode)
                get_zip_income(zipcode)
            # Population within rings (slowest — do last)
            if iso:
                get_population_in_ring(iso, 10)
                get_population_in_ring(iso, 20)
        except Exception as e:
            print(f"Preload error for {clinic.get('address','?')}: {e}")
        finally:
            _preload_status["done"] += 1

    time.sleep(5)  # let app start serving first
    # Only preload clinics not already in disk cache
    to_preload = [cl for cl in CLINICS if f"geo:{cl['address']}" not in _cache]
    print(f"Preloading {len(to_preload)}/{len(CLINICS)} clinics (others cached)")
    _preload_status["total"] = len(to_preload)
    with ThreadPoolExecutor(max_workers=1) as ex:
        ex.map(load_one, to_preload)

    _preload_status["complete"] = True
    print(f"Preload complete: {_preload_status['done']}/{_preload_status['total']} clinics cached")

threading.Thread(target=_preload_worker, daemon=True).start()


# ── Routes ───────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    sites = load_sites()
    clinics = sites.get("clinics", [])
    territories = sites.get("territories", {})
    return render_template("index.html",
                           clinics=clinics,
                           clinics_json=json.dumps(clinics),
                           territories=territories,
                           territories_json=json.dumps(territories))

@app.route("/preload-status")
@login_required
def preload_status():
    return jsonify(_preload_status)

@app.route("/clinic-coords")
@login_required
def clinic_coords():
    """Return geocoded coordinates + cached isochrones for all clinics.
    Called once on page load so the map populates immediately."""
    results = []
    for clinic in CLINICS:
        coords = geocode_address(clinic["address"])
        if not coords:
            continue
        lat, lon = coords
        iso_key = f"iso:{lat:.4f},{lon:.4f}"
        results.append({
            "name": clinic["name"],
            "address": clinic["address"],
            "lat": lat,
            "lon": lon,
            "isochrone": _cache.get(iso_key),  # None if not yet cached
            "color": clinic.get("color", "blue"),
        })
    return jsonify(results)

@app.route("/isochrone", methods=["POST"])
@login_required
def isochrone_endpoint():
    """Fetch isochrone for a single address (used for custom locations)."""
    data = request.get_json()
    address = (data.get("address") or "").strip()
    name = data.get("name", address)
    if not address:
        return jsonify({"error": "No address"}), 400
    # Accept direct lat/lon for partner facilities (skip geocoding)
    direct_lat = data.get("lat")
    direct_lon = data.get("lon")
    if direct_lat and direct_lon:
        lat, lon = float(direct_lat), float(direct_lon)
    else:
        coords = geocode_address(address)
        if not coords:
            return jsonify({"error": f"Could not geocode: {address}"})
        lat, lon = coords
    iso = get_isochrone(lat, lon)
    return jsonify({"name": name, "address": address, "lat": lat, "lon": lon, "isochrone": iso})


@app.route("/compare")
@login_required
def compare_page():
    sites = load_sites()
    territories = sites.get("territories", {})
    # Load Looker productivity data
    looker_data = {}
    try:
        import os as _os
        p2_path = _os.path.join(_os.path.dirname(__file__), "phase2_results.json")
        with open(p2_path) as _f:
            _p2 = json.load(_f)
        for _r in _p2.get("results", []):
            if _r.get("created_initials") is not None:
                looker_data[_r["name"]] = {
                    "created_initials": _r["created_initials"],
                    "fulfilled_initials": _r.get("fulfilled_initials"),
                }
    except Exception as _e:
        print(f"Looker data load error: {_e}")
    return render_template("compare.html", territories=territories,
                           looker_data_json=json.dumps(looker_data))

@app.route("/demographics", methods=["POST"])
@login_required
def demographics():
    """Fetch ESRI (primary) + CMS/SAHIE data for a single address."""
    data = request.get_json()
    address = (data.get("address") or "").strip()
    name = (data.get("name") or "").strip()
    if not address:
        return jsonify({"error": "No address"}), 400
    # Accept direct lat/lon for partner facilities (skip geocoding)
    direct_lat = data.get("lat")
    direct_lon = data.get("lon")
    if direct_lat and direct_lon:
        lat, lon = float(direct_lat), float(direct_lon)
    else:
        coords = geocode_address(address)
        if not coords:
            return jsonify({"error": f"Could not geocode: {address}"})
        lat, lon = coords
    zipcode = extract_zip(address)
    state_fips, county_fips = get_county_fips(lat, lon)

    # ── ESRI: 20-min isochrone polygon enrichment (primary) ─────────────
    from esri_scorer import get_esri_demographics, get_esri_demographics_polygon
    iso_key = f"iso:{lat:.4f},{lon:.4f}"
    cached_iso = _cache.get(iso_key)

    # Run ESRI polygon + CMS/SAHIE concurrently
    from concurrent.futures import ThreadPoolExecutor as _TPE
    def _esri_poly():
        if cached_iso:
            return get_esri_demographics_polygon(lat, lon, cached_iso, minutes=20) or {}
        return {}
    def _esri_point():
        # 1-mile ring fallback for growth rate (polygon call doesn't always include PopulationTotals)
        return get_esri_demographics(lat, lon) or {}
    def _sahie(): return get_sahie(state_fips, county_fips) if state_fips else {}
    def _cms(): return get_cms(zipcode, state_fips=state_fips) if zipcode else {}
    with _TPE(max_workers=4) as ex:
        f_poly  = ex.submit(_esri_poly)
        f_point = ex.submit(_esri_point)
        f_sahie = ex.submit(_sahie)
        f_cms   = ex.submit(_cms)
        esri_poly  = f_poly.result()
        esri_point = f_point.result()
        sahie = f_sahie.result()
        cms   = f_cms.result()

    # Prefer polygon data; fall back to 1-mile ring if polygon not available
    esri = esri_poly if esri_poly.get("total_population") else esri_point
    esri_label = "20-min drive ring" if esri_poly.get("total_population") else "1-mile ring"

    # ── 10-min ring HHI: use cached isochrone + ESRI polygon enrich ──
    hhi_10min = None
    try:
        from esri_scorer import get_esri_demographics_polygon
        cached_iso_raw = _cache.get(f"iso:{lat:.4f},{lon:.4f}")
        if cached_iso_raw:
            iso_10 = get_esri_demographics_polygon(lat, lon, cached_iso_raw, minutes=10)
            if iso_10:
                hhi_10min = iso_10.get('avg_hhi') or iso_10.get('median_income')
    except Exception as _e10:
        print(f"10-min HHI error: {_e10}")

    # ── CTR lookup by clinic name ──
    ctr_pct = CTR_DATA.get(name) if name else None

    # ── Scores ──
    median_hhi_val = esri.get("median_income")
    daytime_pop_val = esri.get("daytime_population")
    base_score     = compute_base_score(median_hhi_val, daytime_pop_val)
    enhanced_score = compute_enhanced_score(median_hhi_val, daytime_pop_val, ctr_pct)

    return jsonify({
        "lat": lat, "lon": lon, "zip": zipcode,
        # ── ESRI demographics ──
        "total_population": esri.get("total_population"),
        "daytime_population": esri.get("daytime_population"),
        "population_45plus": esri.get("population_45plus"),
        "pop45_pct": esri.get("pop45_pct"),
        "median_income": esri.get("median_income"),
        "avg_hhi_10min": hhi_10min,
        "population_growth_pct": esri_point.get("population_growth_pct"),
        "esri_source": bool(esri.get("total_population")),
        "esri_label": esri_label,
        # ── CTR & scores ──
        "ctr_pct": ctr_pct,
        "base_score": base_score,
        "enhanced_score": enhanced_score,
        # ── SAHIE (county insured%) ──
        "insured_pct": sahie.get("insured_pct"),
        # ── CMS procedure volumes ──
        "cpt36475_volume": cms.get("cpt36475_volume"),
        "cpt36465_volume": cms.get("cpt36465_volume"),
        "cpt36466_volume": cms.get("cpt36466_volume"),
        "cpt_total_volume": cms.get("cpt_total_volume"),
        "medicare_rate_36475": cms.get("medicare_rate_36475"),
        "medicare_rate_36465": cms.get("medicare_rate_36465"),
        "medicare_rate_36466": cms.get("medicare_rate_36466"),
        # ── Competitors ──
        "competitors": cms.get("competitors", []),
        "median_income_zip": esri.get("median_income"),
    })



@app.route("/payer-mix", methods=["POST"])
@login_required
def payer_mix_PLACEHOLDER():
    pass


@app.route("/payer-mix", methods=["POST"])
@login_required
def payer_mix():
    """Return Mark Farrah county-level payer mix for a clinic address."""
    data    = request.get_json()
    address = (data.get("address") or "").strip()
    MF_DIR  = '/opt/mikala-apps/clinic-demographics/mark_farrah'
    try:
        fips_cache = json.load(open(f'{MF_DIR}/clinic_fips_cache.json'))
        prod_db    = json.load(open(f'{MF_DIR}/county_product_db.json'))
        carr_db    = json.load(open(f'{MF_DIR}/county_carrier_db.json'))
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    entry = fips_cache.get(address)
    if not entry:
        return jsonify({'error': 'county not found'}), 404
    fips     = str(entry['fips'])
    county   = entry.get('county','')
    state    = entry.get('state','')
    prod     = prod_db.get(fips, {})
    carr_key = f"{state}|{county}"
    carr     = carr_db.get(carr_key, {})
    priv = prod.get('priv_total', 0)
    if not priv:
        return jsonify({'error': 'no data for county'}), 404
    plan_type = [
        {'label': 'PPO',     'val': prod.get('ppo',0),     'pct': prod.get('pct_ppo',0)},
        {'label': 'ASO',     'val': prod.get('aso',0),     'pct': prod.get('pct_aso',0)},
        {'label': 'POS',     'val': prod.get('pos',0),     'pct': round(prod.get('pos',0)/priv*100,1)},
        {'label': 'HMO/EPO', 'val': prod.get('hmo_epo',0), 'pct': prod.get('pct_hmo_epo',0)},
    ]
    carriers = [
        {'label': c[0], 'val': c[1], 'pct': c[2]}
        for c in (carr.get('carriers') or [])[:6]
    ]
    return jsonify({
        'county': county, 'state': state, 'fips': fips,
        'priv_total': priv,
        'addressable': prod.get('addressable',0),
        'pct_addressable': prod.get('pct_addressable',0),
        'plan_type': plan_type,
        'carriers': carriers,
    })


@app.route("/ring-pop", methods=["POST"])
@login_required
def ring_pop():
    """Async endpoint for ring population - called by frontend after fast data loads."""
    data = request.get_json()
    address = (data.get("address") or "").strip()
    direct_lat = data.get("lat")
    direct_lon = data.get("lon")
    if direct_lat and direct_lon:
        lat, lon = float(direct_lat), float(direct_lon)
    elif address:
        coords = geocode_address(address)
        if not coords:
            return jsonify({"error": "geocode failed"})
        lat, lon = coords
    else:
        return jsonify({"error": "No address or coords"}), 400
    iso = get_isochrone(lat, lon)
    if not iso:
        return jsonify({"error": "no isochrone"})
    pop10 = get_population_in_ring(iso, 10)
    pop20 = get_population_in_ring(iso, 20)
    return jsonify({
        "pop_10min": pop10.get("total_pop"),
        "female_35plus_10min": pop10.get("female_35plus"),
        "median_income_zip": zip_income,
        "pop_20min": pop20.get("total_pop"),
        "female_35plus_20min": pop20.get("female_35plus"),
        "median_income_20min": pop20.get("median_income"),  # kept for compat
    })


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    f = request.files.get("file")
    if not f or not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Please upload a .xlsx file"}), 400
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        STATUS_COLORS = {
            "open": "blue",
            "pending opening day": "purple",
            "possible new sites": "orange",
        }
        territories = {}
        clinics = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            territory_clinics = []
            for row in rows[1:]:
                if len(row) < 4:
                    continue
                name, address, status, include = row[0], row[1], row[2], row[3]
                if include is not True and str(include).strip().lower() not in ("true", "yes", "1"):
                    continue
                if not name or not address:
                    continue
                status = str(status).strip() if status else ""
                color = STATUS_COLORS.get(status.lower(), "blue")
                clinic = {
                    "name": str(name).strip(),
                    "address": str(address).strip(),
                    "status": status,
                    "color": color,
                    "territory": sheet_name,
                }
                territory_clinics.append(clinic)
                clinics.append(clinic)
            if territory_clinics:
                territories[sheet_name] = territory_clinics
        data = {"territories": territories, "clinics": clinics}
        with open("/opt/mikala-apps/clinic-demographics/vip_sites.json", "w") as out:
            json.dump(data, out)
        global CLINICS, _preload_status
        new_only = _smart_cache_refresh(CLINICS, clinics)
        CLINICS = clinics
        # Only preload new/changed clinics; existing cache stays warm
        to_preload = new_only if new_only else clinics
        _preload_status = {"done": len(clinics) - len(to_preload), "total": len(clinics), "complete": len(to_preload) == 0}
        import threading as _t
        if to_preload:
            def _partial_preload():
                from concurrent.futures import ThreadPoolExecutor
                def load_one(clinic):
                    try:
                        coords = geocode_address(clinic["address"])
                        if coords:
                            lat, lon = coords
                            iso = get_isochrone(lat, lon)
                            zipcode = extract_zip(clinic["address"])
                            state_fips, county_fips = get_county_fips(lat, lon)
                            if state_fips:
                                get_census_acs(state_fips, county_fips)
                                get_sahie(state_fips, county_fips)
                            if zipcode:
                                get_cms(zipcode)
                            if iso:
                                get_population_in_ring(iso, 10)
                                get_population_in_ring(iso, 20)
                    except Exception as e:
                        print(f"Preload error: {e}")
                    finally:
                        _preload_status["done"] += 1
                with ThreadPoolExecutor(max_workers=3) as ex:
                    ex.map(load_one, to_preload)
                _preload_status["complete"] = True
            _t.Thread(target=_partial_preload, daemon=True).start()
        return jsonify({"ok": True, "count": len(clinics), "territories": list(territories.keys()), "new_clinics": len(new_only)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/sites-data")
@login_required
def sites_data():
    try:
        with open("/opt/mikala-apps/clinic-demographics/vip_sites.json") as f:
            return jsonify(json.load(f))
    except:
        return jsonify({"territories": {}, "clinics": []})


@app.route("/sync", methods=["POST"])
@login_required
def sync_from_sheets():
    try:
        import sys, os
        sys.path.insert(0, os.path.dirname(__file__))
        import sheets_sync
        # Update config path to app directory
        sheets_sync.CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'google_sheets_config.json')
        count, tabs = sheets_sync.sync()
        from sheets_sync import SITES_FILE
        new_clinics = load_sites().get('clinics', [])
        global CLINICS, _preload_status
        new_only = _smart_cache_refresh(CLINICS, new_clinics)
        CLINICS = new_clinics
        to_preload = new_only if new_only else new_clinics
        _preload_status = {'done': len(new_clinics) - len(to_preload), 'total': len(new_clinics), 'complete': len(to_preload) == 0}
        import threading as _t
        if to_preload:
            def _partial_preload_sheets():
                from concurrent.futures import ThreadPoolExecutor
                def load_one(clinic):
                    try:
                        coords = geocode_address(clinic["address"])
                        if coords:
                            lat, lon = coords
                            iso = get_isochrone(lat, lon)
                            zipcode = extract_zip(clinic["address"])
                            state_fips, county_fips = get_county_fips(lat, lon)
                            if state_fips:
                                get_census_acs(state_fips, county_fips)
                                get_sahie(state_fips, county_fips)
                            if zipcode:
                                get_cms(zipcode)
                            if iso:
                                get_population_in_ring(iso, 10)
                                get_population_in_ring(iso, 20)
                    except Exception as e:
                        print(f"Preload error: {e}")
                    finally:
                        _preload_status["done"] += 1
                with ThreadPoolExecutor(max_workers=3) as ex:
                    ex.map(load_one, to_preload)
                _preload_status["complete"] = True
            _t.Thread(target=_partial_preload_sheets, daemon=True).start()
        return jsonify({'ok': True, 'count': count, 'territories': tabs, 'new_clinics': len(new_only)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)


@app.route('/clinic-performance')
@login_required
def clinic_performance_page():
    return render_template('clinic_performance.html')

@app.route('/api/clinic-performance')
@login_required
def api_clinic_performance():
    start = request.args.get('startDate', '')
    end   = request.args.get('endDate', '')
    campaign_ids = [c.strip() for c in request.args.get('campaignIds','').split(',') if c.strip()]
    if not start or not end:
        return jsonify({'error': 'startDate and endDate required'}), 400
    data, err = get_clinic_performance(start, end, campaign_ids or None)
    if err:
        return jsonify({'error': err}), 500
    return jsonify(data)

@app.route('/api/campaigns-list')
@login_required
def api_campaigns_list():
    start = request.args.get('startDate', '')
    end   = request.args.get('endDate', '')
    if not start or not end:
        return jsonify({'error': 'startDate and endDate required'}), 400
    campaigns, err = get_campaigns_list(start, end)
    if err:
        return jsonify({'error': err}), 500
    return jsonify({'campaigns': campaigns})


@app.route('/clinic-demographics/api/patient-heatmap')
@app.route('/api/patient-heatmap')
@login_required
def api_patient_heatmap():
    clinic = request.args.get('clinic', '')
    if not clinic:
        return jsonify({'error': 'clinic required'}), 400
    try:
        points, total = get_heatmap_points(clinic)
        if points is None:
            return jsonify({'error': 'No data for clinic', 'clinic': clinic}), 404
        return jsonify({'points': points, 'total': total, 'clinic': clinic})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/ring-ads', methods=['POST'])
@login_required
def api_ring_ads():
    data = request.get_json()
    lat = data.get('lat')
    lon = data.get('lon')
    start_date = data.get('startDate', '')
    end_date   = data.get('endDate', '')
    if not all([lat, lon, start_date, end_date]):
        return jsonify({'error': 'lat, lon, startDate, endDate required'}), 400
    result = compute_ring_ads(float(lat), float(lon), start_date, end_date)
    return jsonify(result)



# Partner facilities cache (loaded once at startup)
_partner_cache = None
def _get_partner_data():
    global _partner_cache
    if _partner_cache is None:
        import os
        base = '/opt/mikala-apps/clinic-demographics'
        keep = {'name','lat','lng','city','state','type'}
        result = {'uspi': [], 'sca': [], 'nuehealth': []}
        for key, fname in [('uspi','uspi_locations.json'),('sca','sca_locations.json'),('nuehealth','nuehealth_locations.json')]:
            path = os.path.join(base, fname)
            if os.path.exists(path):
                with open(path) as fh:
                    data = json.load(fh)
                result[key] = [{k:v for k,v in f.items() if k in keep}
                               for f in data if f.get('lat') and f.get('lng')]
        _partner_cache = result
    return _partner_cache

@app.route('/clinic-demographics/api/buxton-scores')
@app.route('/api/buxton-scores')
@login_required
def buxton_scores():
    import os as _os
    path = _os.path.join(_os.path.dirname(__file__), 'buxton_scores.json')
    try:
        with open(path) as f2:
            return jsonify(json.load(f2))
    except Exception:
        return jsonify({})

@app.route('/clinic-demographics/api/partner-facilities')
@app.route('/api/partner-facilities')
@login_required
def api_partner_facilities():
    from flask import make_response
    data = _get_partner_data()
    resp = make_response(jsonify(data))
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp



# ── Competitors API ──────────────────────────────────────────────────────────

@app.route('/clinic-demographics/api/competitors', methods=['POST'])
@app.route('/api/competitors', methods=['POST'])
@login_required
def api_competitors():
    """Return T1/T2 classified competitors for a clinic."""
    import os as _os
    data  = request.get_json() or {}
    name  = (data.get('name') or '').strip()
    if not name:
        return jsonify([])

    cache_path = _os.path.join(_os.path.dirname(__file__), 'competitor_cache.json')
    try:
        with open(cache_path) as f:
            comp_cache = json.load(f)
    except Exception:
        return jsonify([])

    entry = comp_cache.get(name)
    if not entry:
        # Try fuzzy match
        for k in comp_cache:
            if k.lower() == name.lower() or name.lower() in k.lower() or k.lower() in name.lower():
                entry = comp_cache[k]
                break
    if not entry:
        return jsonify([])

    places = entry.get('places', [])
    result = []
    seen   = set()
    for p in places:
        lat = p.get('lat')
        lon = p.get('lon')
        if not lat or not lon:
            continue
        key = f"{lat:.4f},{lon:.4f}"
        if key in seen:
            continue
        seen.add(key)
        rating  = p.get('rating') or 0
        reviews = p.get('reviews') or 0
        # Tier classification
        if rating >= 4.5 and reviews >= 20:
            tier = 2
        else:
            tier = 1
        result.append({
            'name':    p.get('name', 'Unknown'),
            'lat':     lat,
            'lon':     lon,
            'rating':  rating,
            'reviews': reviews,
            'tier':    tier,
        })
    # Sort: T2 first, then by review count desc
    result.sort(key=lambda x: (-x['tier'], -x['reviews']))
    return jsonify(result)

# ── Site Scoring Endpoint ─────────────────────────────────────────

@app.route('/clinic-demographics/api/site-score', methods=['POST'])
@login_required
def api_site_score():
    """Compute composite VIP site score for a location."""
    from site_scorer import compute_site_score
    from poi_scorer import get_poi_scores

    data = request.get_json() or {}
    lat = data.get('lat')
    lng = data.get('lng') or data.get('lon')
    address = (data.get('address') or '').strip()

    if not lat or not lng:
        if not address:
            return jsonify({'error': 'lat/lng or address required'}), 400
        coords = geocode_address(address)
        if not coords:
            return jsonify({'error': f'Could not geocode: {address}'}), 400
        lat, lng = coords

    lat, lng = float(lat), float(lng)

    # Cache check (7 days)
    cache_key = f"score:{lat:.3f},{lng:.3f}"
    if cache_key in _cache:
        entry = _cache[cache_key]
        from datetime import datetime, timedelta
        if datetime.utcnow() - entry['_cached_at'] < timedelta(days=7):
            result = dict(entry)
            result.pop('_cached_at', None)
            return jsonify(result)

    # Get demo data if not provided
    demo_data = data.get('demo') or {}
    ring_pop_data = data.get('ring_pop') or {}

    if not demo_data:
        # Try to load from cache or fetch
        if not address:
            # Reverse-geocode not available — use what we have
            pass
        else:
            zipcode = extract_zip(address)
            state_fips, county_fips = get_county_fips(lat, lng)
            if state_fips:
                acs = get_census_acs(state_fips, county_fips)
                sahie = get_sahie(state_fips, county_fips)
                demo_data.update(acs or {})
                demo_data.update(sahie or {})
            if zipcode:
                cms = get_cms(zipcode, state_fips=state_fips)
                demo_data['competitors'] = cms.get('competitors', [])
                zip_inc = get_zip_income(zipcode)
                demo_data['median_income_zip'] = zip_inc

    if not ring_pop_data:
        iso_key = f"iso:{lat:.4f},{lng:.4f}"
        iso = _cache.get(iso_key) or get_isochrone(lat, lng)
        if iso:
            pop10 = get_population_in_ring(iso, 10)
            pop20 = get_population_in_ring(iso, 20)
            ring_pop_data = {
                'pop_10min': pop10.get('total_pop'),
                'pop_20min': pop20.get('total_pop'),
                'female_35plus_10min': pop10.get('female_35plus'),
                'female_35plus_20min': pop20.get('female_35plus'),
            }

    # Compute score
    try:
        score_result = compute_site_score(lat, lng, address, demo_data, ring_pop_data, None)
    except Exception as e:
        return jsonify({'error': f'Scoring error: {e}'}), 500

    # Cache result
    from datetime import datetime
    cached_entry = dict(score_result)
    cached_entry['_cached_at'] = datetime.utcnow()
    _cache[cache_key] = cached_entry

    return jsonify(score_result)
